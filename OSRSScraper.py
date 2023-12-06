import csv
import requests
from openpyxl import load_workbook
#Checks for a valid response
def search_osrs_highscores(username):
    url = f"https://services.runescape.com/m=hiscore_oldschool/index_lite.ws?player={username}"
    response = requests.get(url)
    if response.status_code == 200:
        rows = response.text.strip().split("\n")
        data = [row.split(",")[1:] for row in rows]
        return data
    else:
        print(f"Error: {response.status_code}")
        return None
#Starts the data in a specific cell due to other formulas in the excel document
def export_to_excel(data, filepath):
    workbook = load_workbook(filepath)
    sheet = workbook.active

    headers = ["Level", "Experience"]
    sheet["B4"] = headers[0]
    sheet["C4"] = headers[1]

    for row_num, row_data in enumerate(data, start=5):
        level = int(row_data[0])
        if level <= 0:
            break
        for col_num, cell_value in enumerate(row_data, start=2):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(filepath)
    print(f"Data exported to: {filepath}")

# Search for the specified username on Old School RuneScape high scores. Change the username value and keep the " "!!!
username = "ENTER USERNAME"
highscores_data = search_osrs_highscores(username)

if highscores_data:
    #Change this file path!!!!!!!!!!
    filepath = r"C:\Users\Username\Desktop\MaxSpreadSheet.xlsx"
    export_to_excel(highscores_data, filepath)